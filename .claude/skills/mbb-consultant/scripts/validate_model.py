#!/usr/bin/env python3
"""
Validates an MBB Excel workbook for completeness and correctness.

Usage:
    python validate_model.py input.xlsx

Returns a JSON report with pass/fail status and issues found.
"""

import argparse
import json
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REQUIRED_SHEETS = ["Methodology", "Data Sources", "Calculations", "Summary", "Validation"]

METHODOLOGY_REQUIRED_KEYWORDS = ["objective", "approach", "assumption", "limitation"]

ERROR_FORMULAS = ["#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#N/A", "#VALUE!", "#NUM!"]


def validate_workbook(filepath: str) -> dict:
    """
    Validate an MBB Excel workbook.

    Returns:
        dict with keys: passed (bool), score (int), total_checks (int), issues (list), warnings (list)
    """
    issues = []
    warnings = []
    checks_passed = 0
    total_checks = 0

    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        return {
            "passed": False,
            "score": 0,
            "total_checks": 1,
            "issues": [f"Cannot open workbook: {e}"],
            "warnings": [],
        }

    # ── Check 1: All 5 required sheets present ──────────────────────────────
    total_checks += 1
    sheet_names = wb.sheetnames
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in sheet_names]
    if missing_sheets:
        issues.append(f"Missing required sheets: {', '.join(missing_sheets)}")
    else:
        checks_passed += 1

    # ── Check 2: Methodology sheet has required sections ────────────────────
    total_checks += 1
    if "Methodology" in sheet_names:
        ws = wb["Methodology"]
        all_text = " ".join(
            str(cell.value).lower()
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        )
        missing_keywords = [kw for kw in METHODOLOGY_REQUIRED_KEYWORDS if kw not in all_text]
        if missing_keywords:
            issues.append(f"Methodology sheet missing sections for: {', '.join(missing_keywords)}")
        else:
            checks_passed += 1
    else:
        issues.append("Cannot check Methodology — sheet missing")

    # ── Check 3: Data Sources sheet has entries ─────────────────────────────
    total_checks += 1
    if "Data Sources" in sheet_names:
        ws = wb["Data Sources"]
        data_rows = ws.max_row - 1  # Subtract header
        if data_rows < 1:
            issues.append("Data Sources sheet has no entries (only header row)")
        else:
            checks_passed += 1
    else:
        issues.append("Cannot check Data Sources — sheet missing")

    # ── Check 4: Calculations sheet uses formulas ───────────────────────────
    total_checks += 1
    if "Calculations" in sheet_names:
        ws = wb["Calculations"]
        formula_count = 0
        value_count = 0
        for row in ws.iter_rows(min_row=2):  # Skip header
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
                    elif isinstance(cell.value, (int, float)):
                        value_count += 1

        if formula_count == 0 and value_count > 0:
            issues.append(
                f"Calculations sheet has {value_count} hardcoded values but no formulas. "
                "Use Excel formulas instead of hardcoded Python calculations."
            )
        elif formula_count > 0:
            checks_passed += 1
            ratio = formula_count / (formula_count + value_count) if (formula_count + value_count) > 0 else 0
            if ratio < 0.3:
                warnings.append(
                    f"Low formula ratio ({ratio:.0%}). Consider using more formulas vs. hardcoded values."
                )
        else:
            warnings.append("Calculations sheet appears empty")
            checks_passed += 1  # Empty is not a failure
    else:
        issues.append("Cannot check Calculations — sheet missing")

    # ── Check 5: No formula errors ──────────────────────────────────────────
    total_checks += 1
    error_cells = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value) in ERROR_FORMULAS:
                    error_cells.append(
                        f"{sheet_name}!{get_column_letter(cell.column)}{cell.row}: {cell.value}"
                    )

    if error_cells:
        issues.append(f"Formula errors found: {'; '.join(error_cells[:10])}")
        if len(error_cells) > 10:
            issues.append(f"  ... and {len(error_cells) - 10} more errors")
    else:
        checks_passed += 1

    # ── Check 6: Summary sheet has findings and recommendations ─────────────
    total_checks += 1
    if "Summary" in sheet_names:
        ws = wb["Summary"]
        all_text = " ".join(
            str(cell.value).lower()
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        )
        has_findings = any(kw in all_text for kw in ["finding", "key", "result", "value"])
        has_recommendations = "recommend" in all_text
        if not has_findings:
            warnings.append("Summary sheet may be missing key findings")
        if not has_recommendations:
            warnings.append("Summary sheet may be missing recommendations")
        if has_findings or has_recommendations:
            checks_passed += 1
        else:
            issues.append("Summary sheet appears to have no findings or recommendations")
    else:
        issues.append("Cannot check Summary — sheet missing")

    # ── Check 7: Validation sheet has sensitivity analysis ──────────────────
    total_checks += 1
    if "Validation" in sheet_names:
        ws = wb["Validation"]
        all_text = " ".join(
            str(cell.value).lower()
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        )
        has_sensitivity = "sensitivity" in all_text or "scenario" in all_text
        has_checks = "check" in all_text or "cross" in all_text or "sanity" in all_text
        if not has_sensitivity:
            warnings.append("Validation sheet may be missing sensitivity analysis")
        if not has_checks:
            warnings.append("Validation sheet may be missing cross-checks")
        if has_sensitivity or has_checks:
            checks_passed += 1
        else:
            issues.append("Validation sheet appears to have no sensitivity analysis or cross-checks")
    else:
        issues.append("Cannot check Validation — sheet missing")

    # ── Check 8: Input cells use blue font ──────────────────────────────────
    total_checks += 1
    if "Calculations" in sheet_names:
        ws = wb["Calculations"]
        blue_input_count = 0
        non_blue_input_count = 0
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    fill_color = str(cell.fill.start_color.rgb)
                    if "FFF2CC" in fill_color or "FFFFF2CC" in fill_color:
                        # This is an input cell (yellow background)
                        if cell.font and cell.font.color and cell.font.color.rgb:
                            font_color = str(cell.font.color.rgb)
                            if "0000FF" in font_color or "000000FF" in font_color:
                                blue_input_count += 1
                            else:
                                non_blue_input_count += 1

        if non_blue_input_count > 0 and blue_input_count == 0:
            warnings.append(
                f"Input cells (yellow background) should use blue font. "
                f"Found {non_blue_input_count} input cells without blue font."
            )
        else:
            checks_passed += 1
    else:
        checks_passed += 1  # Can't check, not a failure

    # ── Result ──────────────────────────────────────────────────────────────
    passed = len(issues) == 0
    return {
        "passed": passed,
        "score": checks_passed,
        "total_checks": total_checks,
        "issues": issues,
        "warnings": warnings,
    }


def main():
    parser = argparse.ArgumentParser(description="Validate MBB Excel workbook")
    parser.add_argument("filepath", help="Path to Excel workbook")
    parser.add_argument("--format", choices=["json", "text"], default="text",
                        help="Output format")
    args = parser.parse_args()

    result = validate_workbook(args.filepath)

    if args.format == "json":
        print(json.dumps(result, indent=2))
    else:
        status = "PASSED" if result["passed"] else "FAILED"
        print(f"\nValidation {status} ({result['score']}/{result['total_checks']} checks passed)")
        print("=" * 60)

        if result["issues"]:
            print("\nISSUES (must fix):")
            for issue in result["issues"]:
                print(f"  [X] {issue}")

        if result["warnings"]:
            print("\nWARNINGS (review):")
            for warning in result["warnings"]:
                print(f"  [!] {warning}")

        if not result["issues"] and not result["warnings"]:
            print("\nAll checks passed. Workbook meets MBB standards.")

    sys.exit(0 if result["passed"] else 1)


if __name__ == "__main__":
    main()
