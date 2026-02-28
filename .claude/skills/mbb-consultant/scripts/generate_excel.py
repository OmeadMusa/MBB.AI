#!/usr/bin/env python3
"""
MBB-style Excel workbook generator.

Usage (CLI):
    python generate_excel.py --output FILE --title TITLE [options]

Usage (Python):
    from generate_excel import MBBWorkbookBuilder
    builder = MBBWorkbookBuilder("Analysis Title")
    builder.add_methodology(...)
    builder.add_data_source(...)
    builder.add_calculation_headers(...)
    builder.add_calculation_row(...)
    builder.add_summary(...)
    builder.add_sensitivity(...)
    builder.add_cross_check(...)
    builder.save("output.xlsx")
"""

from __future__ import annotations

import argparse
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter


# ── MBB Color Constants ──────────────────────────────────────────────────────

DARK_BLUE = "1F4E79"
MEDIUM_BLUE = "2E75B6"
LIGHT_GRAY = "F2F2F2"
INPUT_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
BLACK = "000000"
BLUE_FONT = "0000FF"
GREEN_FONT = "008000"

HEADER_FILL = PatternFill(start_color=DARK_BLUE, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, fill_type="solid")
INPUT_FILL = PatternFill(start_color=INPUT_YELLOW, fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
SUBHEADER_FONT = Font(name="Calibri", size=11, color=WHITE)
BODY_FONT = Font(name="Calibri", size=11, color=BLACK)
BOLD_FONT = Font(name="Calibri", size=11, bold=True, color=BLACK)
INPUT_FONT = Font(name="Calibri", size=11, color=BLUE_FONT)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
XREF_FONT = Font(name="Calibri", size=11, color=GREEN_FONT)

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)
TOTAL_BORDER = Border(
    top=Side(style="medium", color=BLACK),
    bottom=Side(style="double", color=BLACK),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")


class MBBWorkbookBuilder:
    """Creates MBB-standard Excel workbooks with 5 required sheets."""

    SHEET_NAMES = ["Methodology", "Data Sources", "Calculations", "Summary", "Validation"]

    def __init__(self, title: str):
        self.title = title
        self.wb = Workbook()
        self.date = datetime.now().strftime("%B %d, %Y")

        # Create all 5 sheets
        self.wb.active.title = self.SHEET_NAMES[0]
        for name in self.SHEET_NAMES[1:]:
            self.wb.create_sheet(name)

        self._calc_row = 1  # Track current row in Calculations sheet
        self._source_row = 2  # Track current row in Data Sources sheet
        self._sensitivity_row = 1  # Track current row in Validation sheet
        self._cross_check_row = None  # Set after sensitivity section

        self._init_sheets()

    def _init_sheets(self):
        """Initialize all sheets with headers and base formatting."""
        self._init_methodology()
        self._init_data_sources()
        self._init_summary()
        self._init_validation()

    def _apply_header_row(self, ws, headers, row=1):
        """Apply MBB header formatting to a row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
        ws.row_dimensions[row].height = 25

    def _auto_fit_columns(self, ws, min_width=12):
        """Auto-fit column widths based on content."""
        for col_idx in range(1, ws.max_column + 1):
            max_len = min_width
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)) + 2)
            ws.column_dimensions[col_letter].width = min(max_len, 50)

    # ── Methodology Sheet ────────────────────────────────────────────────────

    def _init_methodology(self):
        ws = self.wb["Methodology"]
        ws.cell(row=1, column=1, value="Methodology & Approach").font = TITLE_FONT
        ws.cell(row=1, column=1).alignment = LEFT_ALIGN
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 30

    def add_methodology(self, objective: str, approach: str,
                        assumptions: list[str], limitations: list[str]):
        """Populate the Methodology sheet."""
        ws = self.wb["Methodology"]
        row = 3

        sections = [
            ("Analysis Objective", objective),
            ("Analytical Approach", approach),
            ("Date of Analysis", self.date),
        ]

        for label, value in sections:
            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            ws.cell(row=row, column=2, value=value).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = LEFT_ALIGN
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Key Assumptions").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=row, column=1).font = SUBHEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        for i, assumption in enumerate(assumptions, 1):
            ws.cell(row=row, column=1, value=f"{i}.").font = BODY_FONT
            ws.cell(row=row, column=1).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=2, value=assumption).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = LEFT_ALIGN
            if i % 2 == 0:
                ws.cell(row=row, column=1).fill = ALT_ROW_FILL
                ws.cell(row=row, column=2).fill = ALT_ROW_FILL
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Limitations & Caveats").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=row, column=1).font = SUBHEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        for i, limitation in enumerate(limitations, 1):
            ws.cell(row=row, column=1, value=f"{i}.").font = BODY_FONT
            ws.cell(row=row, column=1).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=2, value=limitation).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = LEFT_ALIGN
            if i % 2 == 0:
                ws.cell(row=row, column=1).fill = ALT_ROW_FILL
                ws.cell(row=row, column=2).fill = ALT_ROW_FILL
            row += 1

        self._auto_fit_columns(ws)

    # ── Data Sources Sheet ───────────────────────────────────────────────────

    def _init_data_sources(self):
        ws = self.wb["Data Sources"]
        headers = ["Source Name", "Source Type", "Date Accessed",
                    "Reliability", "Reference", "Notes"]
        self._apply_header_row(ws, headers)

    def add_data_source(self, name: str, source_type: str, date: str,
                        reliability: str, reference: str, notes: str = ""):
        """Add a row to the Data Sources sheet."""
        ws = self.wb["Data Sources"]
        row = self._source_row
        values = [name, source_type, date, reliability, reference, notes]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = BODY_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL

        # Color-code reliability
        rel_cell = ws.cell(row=row, column=4)
        if reliability.lower() == "high":
            rel_cell.font = Font(name="Calibri", size=11, color="008000")
        elif reliability.lower() == "low":
            rel_cell.font = Font(name="Calibri", size=11, color="FF0000")

        ws.row_dimensions[row].height = 18
        self._source_row += 1
        self._auto_fit_columns(ws)

    # ── Calculations Sheet ───────────────────────────────────────────────────

    def add_calculation_headers(self, headers: list[str], units: list[str] | None = None):
        """Set up the Calculations sheet header row with optional unit sub-headers."""
        ws = self.wb["Calculations"]
        self._apply_header_row(ws, headers, row=1)
        self._calc_row = 2

        if units:
            for col, unit in enumerate(units, 1):
                if unit:
                    cell = ws.cell(row=2, column=col, value=f"({unit})")
                    cell.font = SUBHEADER_FONT
                    cell.fill = SUBHEADER_FILL
                    cell.alignment = CENTER_ALIGN
            ws.row_dimensions[2].height = 20
            self._calc_row = 3

    def add_calculation_row(self, data: list, row_type: str = "data"):
        """
        Add a row to the Calculations sheet.

        Args:
            data: List of values. Strings starting with '=' are treated as formulas.
            row_type: One of 'data', 'input', 'formula', 'total', 'blank'
        """
        ws = self.wb["Calculations"]
        row = self._calc_row

        if row_type == "blank":
            self._calc_row += 1
            return

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)

            if value is None:
                cell.value = None
            elif isinstance(value, str) and value.startswith("="):
                cell.value = value
            else:
                cell.value = value

            # Apply formatting based on row_type
            if row_type == "input":
                cell.font = INPUT_FONT
                cell.fill = INPUT_FILL
            elif row_type == "total":
                cell.font = BOLD_FONT
                cell.border = TOTAL_BORDER
            elif row_type == "formula":
                cell.font = BODY_FONT
            else:
                cell.font = BODY_FONT

            # Alternate row shading for data rows
            if row_type == "data" and row % 2 == 0:
                cell.fill = ALT_ROW_FILL

            cell.alignment = RIGHT_ALIGN if col > 1 else LEFT_ALIGN
            if row_type != "total":
                cell.border = THIN_BORDER

        ws.row_dimensions[row].height = 18
        self._calc_row += 1

    def add_calculation_section(self, section_title: str):
        """Add a section sub-header in the Calculations sheet."""
        ws = self.wb["Calculations"]
        row = self._calc_row
        cell = ws.cell(row=row, column=1, value=section_title)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ws.max_column or 5)
        ws.row_dimensions[row].height = 22
        self._calc_row += 1

    # ── Summary Sheet ────────────────────────────────────────────────────────

    def _init_summary(self):
        ws = self.wb["Summary"]
        ws.cell(row=1, column=1, value=f"Executive Summary: {self.title}").font = TITLE_FONT
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 30

    def add_summary(self, findings: list[tuple], recommendations: list[str]):
        """
        Populate the Summary sheet.

        Args:
            findings: List of tuples (finding, value, benchmark, implication)
            recommendations: List of recommendation strings
        """
        ws = self.wb["Summary"]
        row = 3

        # Key Findings table
        finding_headers = ["Key Finding", "Value", "vs. Benchmark", "Implication"]
        self._apply_header_row(ws, finding_headers, row=row)
        row += 1

        for i, (finding, value, benchmark, implication) in enumerate(findings):
            data = [finding, value, benchmark, implication]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = BODY_FONT
                cell.alignment = LEFT_ALIGN
                cell.border = THIN_BORDER
                if i % 2 == 1:
                    cell.fill = ALT_ROW_FILL
            ws.row_dimensions[row].height = 18
            row += 1

        row += 2

        # Recommendations section
        ws.cell(row=row, column=1, value="Recommendations").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=row, column=1).font = SUBHEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        for i, rec in enumerate(recommendations, 1):
            ws.cell(row=row, column=1, value=f"{i}.").font = BOLD_FONT
            ws.cell(row=row, column=1).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=2, value=rec).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = LEFT_ALIGN
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            if i % 2 == 0:
                for c in range(1, 5):
                    ws.cell(row=row, column=c).fill = ALT_ROW_FILL
            row += 1

        self._auto_fit_columns(ws)

    # ── Validation Sheet ─────────────────────────────────────────────────────

    def _init_validation(self):
        ws = self.wb["Validation"]
        ws.cell(row=1, column=1, value="Validation & Quality Checks").font = TITLE_FONT
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 30
        self._sensitivity_row = 3

    def add_sensitivity(self, variable_name: str, base_value, variations: list):
        """
        Add a sensitivity analysis table to the Validation sheet.

        Args:
            variable_name: Name of the variable being varied
            base_value: The base case value
            variations: List of alternative values to test
        """
        ws = self.wb["Validation"]
        row = self._sensitivity_row

        # Section header
        ws.cell(row=row, column=1, value=f"Sensitivity: {variable_name}").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=row, column=1).font = SUBHEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(variations) + 1)
        row += 1

        # Headers
        headers = [variable_name] + [str(v) for v in variations]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
        row += 1

        # Mark base case
        for col, v in enumerate(variations, 2):
            cell = ws.cell(row=row, column=col, value="Base" if v == base_value else "")
            cell.font = BODY_FONT
            cell.alignment = CENTER_ALIGN
            if v == base_value:
                cell.font = Font(name="Calibri", size=11, bold=True, color="008000")
        ws.cell(row=row, column=1, value="Case").font = BOLD_FONT
        row += 1

        # Leave rows for user to fill impact values
        for label in ["Output Impact", "% Change from Base"]:
            ws.cell(row=row, column=1, value=label).font = BODY_FONT
            for col in range(2, len(variations) + 2):
                cell = ws.cell(row=row, column=col)
                cell.fill = INPUT_FILL
                cell.font = INPUT_FONT
                cell.alignment = RIGHT_ALIGN
            row += 1

        self._sensitivity_row = row + 1
        self._cross_check_row = row + 2
        self._auto_fit_columns(ws)

    def add_cross_check(self, description: str, expected: str, actual: str, status: str):
        """Add a cross-check row to the Validation sheet."""
        ws = self.wb["Validation"]

        # Initialize cross-check section if needed
        if self._cross_check_row is None:
            self._cross_check_row = self._sensitivity_row + 1

        row = self._cross_check_row

        # Add section header if this is the first cross-check
        if row == self._sensitivity_row + 1 or (
            ws.cell(row=row - 1, column=1).value != "Cross-Check" and
            not any(ws.cell(row=r, column=1).value == "Cross-Checks & Sanity Tests"
                    for r in range(1, row))
        ):
            ws.cell(row=row, column=1, value="Cross-Checks & Sanity Tests").font = BOLD_FONT
            ws.cell(row=row, column=1).fill = SUBHEADER_FILL
            ws.cell(row=row, column=1).font = SUBHEADER_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1

            headers = ["Description", "Expected", "Actual", "Status"]
            self._apply_header_row(ws, headers, row=row)
            row += 1

        values = [description, expected, actual, status]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = BODY_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

        # Color-code status
        status_cell = ws.cell(row=row, column=4)
        if status.upper() == "PASS":
            status_cell.font = Font(name="Calibri", size=11, bold=True, color="008000")
        elif status.upper() == "FAIL":
            status_cell.font = Font(name="Calibri", size=11, bold=True, color="FF0000")
        else:
            status_cell.font = Font(name="Calibri", size=11, bold=True, color="FF8C00")

        self._cross_check_row = row + 1
        self._auto_fit_columns(ws)

    # ── Save ─────────────────────────────────────────────────────────────────

    def save(self, filepath: str):
        """Auto-fit columns, set print areas, freeze panes, and save."""
        os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else ".", exist_ok=True)

        for sheet_name in self.SHEET_NAMES:
            ws = self.wb[sheet_name]
            self._auto_fit_columns(ws)

            # Freeze panes: freeze below header row, after first column
            if sheet_name in ("Data Sources", "Calculations"):
                ws.freeze_panes = "B2"
            elif sheet_name == "Summary":
                ws.freeze_panes = "A2"

            # Set print area
            if ws.max_row and ws.max_column:
                max_col_letter = get_column_letter(ws.max_column)
                ws.print_area = f"A1:{max_col_letter}{ws.max_row}"

        self.wb.save(filepath)
        print(f"Workbook saved: {filepath}")
        return filepath


# ── CLI Interface ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate MBB-style Excel workbook")
    parser.add_argument("--output", required=True, help="Output file path")
    parser.add_argument("--title", required=True, help="Analysis title")
    parser.add_argument("--methodology", default=None, help="Analytical approach description")
    parser.add_argument("--assumptions", default=None, help="JSON array of assumptions")
    args = parser.parse_args()

    builder = MBBWorkbookBuilder(args.title)

    if args.methodology:
        assumptions = json.loads(args.assumptions) if args.assumptions else []
        builder.add_methodology(
            objective=args.title,
            approach=args.methodology,
            assumptions=assumptions,
            limitations=["See analysis for details"]
        )

    builder.save(args.output)


if __name__ == "__main__":
    main()
